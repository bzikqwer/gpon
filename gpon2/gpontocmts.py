adr = input('Введите папку: ')
file = input('Введите название файла: ')
file2='insert.sql'
# str=int(input('введите количество строк в файле: '))


with open((adr+'\\'+file2),'w') as fl:
    print('truncate table BILLING_DB3.ALMATV_CHANGE_GPON_LOGIN_TMP;',file=fl)
    from openpyxl import *
    wb = load_workbook(adr + '\\' + file)
    sheet = wb['Лист1']
    for i in range(2, 200): # максимальное количество строк
         print('insert into BILLING_DB3.ALMATV_CHANGE_GPON_LOGIN_TMP (id_service_inst, v_ext_ident, v_gpon_login) values('+str(sheet.cell(row=i, column=2).value)+','+'\''+str(sheet.cell(row=i, column=1).value)+'\''+','+'\''+str(sheet.cell(row=i, column=4).value)+'\');',file=fl)
